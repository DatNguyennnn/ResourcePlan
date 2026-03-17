from fastapi import APIRouter, Depends, UploadFile, File
from sqlalchemy.orm import Session
from datetime import datetime
import unicodedata
import openpyxl
import io
import logging
from app.database import get_db
from app.models.employee import Employee
from app.models.project import Project
from app.models.resource_allocation import ResourceAllocation

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/import", tags=["import"])


def find_sheet(wb, keyword: str):
    """Find sheet by normalized keyword match."""
    keyword_norm = unicodedata.normalize("NFC", keyword.lower())
    for name in wb.sheetnames:
        name_norm = unicodedata.normalize("NFC", name.lower())
        if keyword_norm == name_norm:
            return wb[name]
    # Partial match
    for name in wb.sheetnames:
        name_norm = unicodedata.normalize("NFC", name.lower())
        if keyword_norm in name_norm or name_norm in keyword_norm:
            return wb[name]
    return None


@router.post("/excel")
def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    contents = file.file.read()
    wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)

    stats = {"employees": 0, "projects": 0, "allocations": 0}
    sheets_found = []
    logger.info(f"Sheet names: {wb.sheetnames}")

    # Import employees
    ws = find_sheet(wb, "DS Nhân viên")
    if ws:
        sheets_found.append(ws.title)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            emp_id = str(row[1]).strip()
            existing = db.query(Employee).filter(Employee.employee_id == emp_id).first()
            if not existing:
                emp = Employee(
                    employee_id=emp_id,
                    full_name=str(row[2] or "").strip(),
                    department=str(row[3] or "").strip(),
                    level=str(row[4] or "Assessed").strip(),
                    status=str(row[5] or "").strip(),
                )
                db.add(emp)
                stats["employees"] += 1
        db.flush()

    # Import projects
    ws = find_sheet(wb, "DS Dự án")
    if ws:
        sheets_found.append(ws.title)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            code = str(row[1]).strip()
            existing = db.query(Project).filter(Project.project_code == code).first()
            if not existing:
                proj = Project(
                    project_code=code,
                    project_name=str(row[2] or code).strip(),
                    description=str(row[3] or "").strip() if row[3] else None,
                    pm=str(row[4] or "").strip() if row[4] else None,
                    status=str(row[5] or "").strip(),
                )
                db.add(proj)
                stats["projects"] += 1
        db.flush()

    # Import resource allocations
    ws = find_sheet(wb, "Resource Plan")
    if ws:
        sheets_found.append(ws.title)
        # Read header row to get week dates
        header = [cell.value for cell in ws[1]]
        week_columns: dict[int, object] = {}
        for i, val in enumerate(header):
            if isinstance(val, datetime):
                week_columns[i] = val.date()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            emp_id_str = str(row[0]).strip()
            project_code = str(row[4] or "").strip()
            if not project_code:
                continue

            emp = db.query(Employee).filter(Employee.employee_id == emp_id_str).first()
            proj = db.query(Project).filter(Project.project_code == project_code).first()
            if not emp or not proj:
                continue

            for col_idx, week_date in week_columns.items():
                if col_idx < len(row) and row[col_idx] is not None:
                    try:
                        pct = float(row[col_idx])
                        if pct > 0:
                            existing = db.query(ResourceAllocation).filter(
                                ResourceAllocation.employee_id == emp.id,
                                ResourceAllocation.project_id == proj.id,
                                ResourceAllocation.week_start == week_date,
                            ).first()
                            if not existing:
                                alloc = ResourceAllocation(
                                    employee_id=emp.id,
                                    project_id=proj.id,
                                    week_start=week_date,
                                    allocation_percentage=pct,
                                )
                                db.add(alloc)
                                stats["allocations"] += 1
                    except (ValueError, TypeError):
                        pass

    db.commit()
    logger.info(f"Import stats: {stats}, sheets found: {sheets_found}")
    return {"message": "Import successful", "stats": stats, "sheets_found": sheets_found}
